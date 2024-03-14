import io
import asyncio

import pytest
import openpyxl

from src.apps.statistics.models import StatisticData
from src.apps.statistics.data.dashboard import (
    get_data_for_cec_user, 
    get_data_for_region_user, 
    get_data_for_district_user
    )


@pytest.mark.parametrize(
    "batch_times",
    [(0), (1), (5), (10)],
)
@pytest.mark.django_db(transaction=True)
def test_file_cec(batch_times, cec_client, url_name, inventory_unit_factory):
    for _ in range(batch_times):
        inventory_unit = inventory_unit_factory.create()

    data = asyncio.run(get_data_for_cec_user())
    StatisticData.objects.update_or_create(
        name='Dashboard_data_for_CEC',
        defaults={'data': data},
    )
    
    url = url_name(__file__)
    response = cec_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws["B5"].value == "Xududlar nomi"
    if ws["B8"].value != "Jami:":
        assert ws["B8"].value == inventory_unit.district.region.name_uz.capitalize()
        assert ws["C6"].value == inventory_unit.product.name_uz


@pytest.mark.parametrize(
    "for_loop_times",
    [(0), (1), (5), (10)],
)
@pytest.mark.django_db(transaction=True)
def test_file_region(
    for_loop_times,
    region_client,
    url_name,
    inventory_unit_factory,
    responsible_person_factory,
    district_factory,
    region_user,
):
    responsible_persons_data = []
    region = region_user.region
    district = district_factory.create(region=region)
    inventory_unit = None

    for i in range(for_loop_times):
        responsible_person = responsible_person_factory.create(district=district)
        gathered_name = (
            responsible_person.first_name
            + " "
            + responsible_person.last_name
            + " "
            + responsible_person.middle_name
            + " ("
            + responsible_person.passport_serial
            + ")"
        )
        responsible_persons_data.append(gathered_name)
        inventory_unit = inventory_unit_factory.create(district=district)
        inventory_unit.product.category.name_uz += str(i)
        inventory_unit.product.category.save()

    responsible_persons_data.reverse()
    gathered_name = ", \n".join(responsible_persons_data)

    data = asyncio.run(get_data_for_region_user(region.id))
    StatisticData.objects.update_or_create(
        name=f"Dashboard_data_for_region_{region.id}",
        defaults={"data": data},
    )

    url = url_name(__file__)
    response = region_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws["B5"].value == "Tuman/Shahar nomi"

    if gathered_name:
        assert ws["C8"].value == gathered_name
    else:
        assert ws["C8"].value == None

    if ws["B8"].value != "Jami:" and inventory_unit:
        assert ws["B8"].value == inventory_unit.district.name_uz
        assert ws["D6"].value == inventory_unit.product.name_uz


@pytest.mark.parametrize(
    "for_loop_times",
    [(0), (1), (5), (10)],
)
@pytest.mark.django_db(transaction=True)
def test_file_district(
    for_loop_times,
    district_client,
    url_name,
    inventory_unit_factory,
    responsible_person_factory,
    district_user,
    storage_place_factory,
):
    district = district_user.district
    address = storage_place_factory.create(district=district)
    responsible_persons_data = []
    inventory_unit = None

    for i in range(for_loop_times):
        responsible_person = responsible_person_factory.create(district=district)
        gathered_name = (
            responsible_person.first_name
            + " "
            + responsible_person.last_name
            + " "
            + responsible_person.middle_name
            + " ("
            + responsible_person.passport_serial
            + ")"
        )
        responsible_persons_data.append(gathered_name)
        inventory_unit = inventory_unit_factory.create(
            storage_place=address, district=district
        )
        inventory_unit.product.category.name_uz += str(i)
        inventory_unit.product.category.save()
        data = asyncio.run(get_data_for_district_user(district.id))
        StatisticData.objects.update_or_create(
            name=f"Dashboard_data_for_district_{district.id}",
            defaults={"data": data},
        )

    url = url_name(__file__)
    response = district_client.post(url)
    responsible_persons_data.reverse()
    gathered_name = ", \n".join(responsible_persons_data)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws["B5"].value == "Ma'sul shaxs\n(F.I.Sh. passport yoki\nID-karta raqami)"
    if gathered_name:
        assert ws["B8"].value == gathered_name
    else:
        assert ws["B8"].value == None
    if ws["B8"].value != "Jami:" and inventory_unit:
        assert ws["C8"].value == inventory_unit.storage_place.address
        assert ws["D5"].value == inventory_unit.product.category.name_uz
        assert ws["D6"].value == inventory_unit.product.name_uz
