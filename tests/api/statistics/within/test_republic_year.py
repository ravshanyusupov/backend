import io
import asyncio

import pytest
import openpyxl

from src.apps.statistics.models import StatisticData
from src.apps.statistics.data.within import get_data_for_cec_by_year


@pytest.mark.django_db(transaction=True)
def test_republic_year(cec_client, url_name, inventory_unit_factory):
    inventory_unit = inventory_unit_factory.create()

    data = asyncio.run(get_data_for_cec_by_year())
    StatisticData.objects.update_or_create(
        name="Data_for_within_republic_year",
        defaults={"data": data},
    )

    url = url_name(__file__)
    response = cec_client.post(url)
    
    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['A4'].value == "Jixozlar nomi"
    assert ws['B4'].value == "  yil   "
    assert ws['C4'].value == "narxi"
    if ws['D4'].value != "JAMI:":
        assert ws['D4'].value == inventory_unit.district.region.name_uz.capitalize()
        assert ws['A6'].value == inventory_unit.product.category.name_uz
        assert ws['A7'].value == inventory_unit.product.name_uz
        


@pytest.mark.django_db
def test_republic_year_without_data(cec_client, url_name):
    url = url_name(__file__)
    response = cec_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active
    
    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['D4'].value == None