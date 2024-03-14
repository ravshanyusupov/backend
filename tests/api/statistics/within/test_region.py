import io
import asyncio

import pytest
import openpyxl

from src.apps.statistics.models import StatisticData
from src.apps.statistics.data.within import get_data_within_region


@pytest.mark.django_db(transaction=True)
def test_region(region_client, url_name, inventory_unit_factory, region_user, district_factory):
    region = region_user.region
    district = district_factory.create(region=region)
    inventory_unit = inventory_unit_factory.create(district=district)

    data = asyncio.run(get_data_within_region(region.id))
    StatisticData.objects.update_or_create(
        name=f"Data_for_within_region_{region.id}",
        defaults={"data": data},
    )

    url = url_name(__file__)
    response = region_client.post(url)
    
    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B5'].value == "Tuman nomi*"
    if ws['B8'].value != "Jami:":
        assert ws['B8'].value == inventory_unit.district.name_uz
        assert ws['C6'].value == inventory_unit.product.name_uz


@pytest.mark.django_db
def test_region_without_data(region_client, url_name):
    url = url_name(__file__)
    response = region_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active
    
    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B8'].value == "Jami:"