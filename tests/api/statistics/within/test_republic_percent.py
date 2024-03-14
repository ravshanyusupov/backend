import io
import asyncio

import pytest
import openpyxl

from src.apps.statistics.models import StatisticData
from src.apps.statistics.data.within import get_data_for_cec_by_percent


@pytest.mark.django_db(transaction=True)
def test_republic_percent(cec_client, url_name, inventory_unit_factory):
    inventory_unit = inventory_unit_factory.create()

    data = asyncio.run(get_data_for_cec_by_percent())
    StatisticData.objects.update_or_create(
        name="Data_for_within_republic_percent",
        defaults={"data": data},
    )

    url = url_name(__file__)
    response = cec_client.post(url)
    
    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B5'].value == "Xududlar nomi"
    assert ws['C5'].value == "Uchastkalar\n soni"
    if ws['B9'].value != "Jami:":
        assert ws['B8'].value == inventory_unit.district.region.name_uz.capitalize()
        assert ws['D6'].value == inventory_unit.product.name_uz
        assert ws['D7'].value == "soni" and ws['E7'].value == " %  "
        


@pytest.mark.django_db
def test_republic_percent_without_data(cec_client, url_name):
    url = url_name(__file__)
    response = cec_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active
    
    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B9'].value == "Jami: "