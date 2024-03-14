import io
import asyncio

import pytest
import openpyxl

from src.apps.statistics.models import StatisticData
from src.apps.statistics.data.within import get_data_for_cec_by_products


@pytest.mark.django_db(transaction=True)
def test_republic_products(cec_client, url_name, inventory_unit_factory):
    inventory_unit = inventory_unit_factory.create()

    data = asyncio.run(get_data_for_cec_by_products())
    StatisticData.objects.update_or_create(
        name="Data_for_within_republic_product",
        defaults={"data": data},
    )

    url = url_name(__file__)
    response = cec_client.post(url)
    
    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B5'].value == "Xududlar nomi"
    if ws['B8'].value != "Jami:":
        assert ws['B8'].value == inventory_unit.district.region.name_uz.capitalize()
        assert ws['C6'].value == inventory_unit.product.name_uz
        


@pytest.mark.django_db
def test_republic_products_without_data(cec_client, url_name):
    url = url_name(__file__)
    response = cec_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active
    
    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B8'].value == "Jami:"