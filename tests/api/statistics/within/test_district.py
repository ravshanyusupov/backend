import io
import asyncio

import pytest
import openpyxl

from src.apps.statistics.models import StatisticData
from src.apps.statistics.data.within import get_data_within_district


@pytest.mark.django_db(transaction=True)
def test_district(district_client, url_name, inventory_unit_factory, storage_place_factory, district_user):
    district = district_user.district
    address = storage_place_factory.create(district=district)
    inventory_unit = inventory_unit_factory.create(district=district, storage_place=address)

    data = asyncio.run(get_data_within_district(district.id))
    StatisticData.objects.update_or_create(
        name=f"Data_for_within_district_{district.id}",
        defaults={"data": data},
    )

    url = url_name(__file__)
    response = district_client.post(url)
    
    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active

    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B5'].value == "Saqlash joyi\n(manzili)"
    if ws['B8'].value != "Jami:":
        assert ws['B8'].value == inventory_unit.storage_place.address
        assert ws['E6'].value == inventory_unit.product.name_uz


@pytest.mark.django_db
def test_district_without_data(district_client, url_name):
    url = url_name(__file__)
    response = district_client.post(url)

    wb = openpyxl.load_workbook(filename=io.BytesIO(response.content), read_only=True)
    ws = wb.active
    
    assert response.status_code == 200
    assert response.headers["Content-Type"] == "application/ms-excel"
    assert ws['B8'].value == None